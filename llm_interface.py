import gradio as gr
from openai import OpenAI
import requests
import logging
import os
import re
from PIL import Image
from config import config
from comfyui_api import comfyui_api
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from tqdm import tqdm
import json
import random
import queue
import threading
import html
import time
import gc

# 创建一个队列来存储日志消息
log_queue = queue.Queue()

class GradioHandler(logging.Handler):
    def emit(self, record):
        log_entry = self.format(record)
        log_queue.put(log_entry)

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
gradio_handler = GradioHandler()
gradio_handler.setLevel(logging.INFO)
logging.getLogger().addHandler(gradio_handler)

# 预定义的参数模板
PARAMETER_TEMPLATES = {
    "视角": ["ultra wide angle", "close-up", "aerial view", "isometric"],
    "风格": ["photorealistic", "cinematic", "anime style", "oil painting", "comic"],
    "色彩": ["vibrant colors", "muted tones", "monochromatic", "pastel colors"],
    "细节": ["highly detailed", "minimalist", "intricate", "abstract"],
}

# 场景类型列表
SCENE_TYPES = [
    "自然风景", "人物特写", "人物中景", "人物远景", "城市风景",
    "魔幻场景", "科幻场景", "平面设计", "LOGO", "产品海报"
]

def load_prompt_templates(file_path='prompt_templates.json'):
    with open(file_path, 'r', encoding='utf-8') as file:
        return json.load(file)

def analyze_prompt_templates(api_key, llm_type, llm_endpoint, llm_model):
    templates = load_prompt_templates()
    analysis_results = {}

    for scene_type, prompts in templates.items():
        prompt = f"""Analyze the following prompts for {scene_type} scenes:

{json.dumps(prompts, indent=2)}

Summarize the key characteristics and patterns in these prompts. Focus on:
1. Common themes and elements
2. Structure and flow of the descriptions
3. Use of specific language or terminology
4. Techniques for creating vivid imagery

Provide a concise summary that can be used as a guide for generating similar high-quality prompts."""

        response = call_llm(api_key, llm_type, llm_endpoint, llm_model, prompt)
        analysis_results[scene_type] = response

    return analysis_results

def enhance_prompt_with_template(api_key, llm_type, llm_endpoint, llm_model, scene_type, keyword, description, generated_prompt):
    templates = load_prompt_templates()
    example_prompts = templates.get(scene_type, [])
    
    if not example_prompts:
        return generated_prompt

    prompt = f"""Given the following information:
1. Scene type: {scene_type}
2. Keyword: {keyword}
3. Description: {description}
4. Initially generated prompt: {generated_prompt}

And these example prompts for {scene_type}:
{json.dumps(example_prompts, indent=2)}

Please enhance the initially generated prompt by incorporating elements and styles from the example prompts. The enhanced prompt should:
1. Maintain the core idea from the original keyword and description
2. Adopt the structure and flow similar to the example prompts
3. Include vivid and specific details that are characteristic of {scene_type} scenes
4. Use language and terminology that creates a strong visual image

Provide only the enhanced prompt as your response, without any additional explanations."""

    enhanced_prompt = call_llm(api_key, llm_type, llm_endpoint, llm_model, prompt)
    return enhanced_prompt.strip()

def generate_prompt_without_optimization(api_key, description, llm_type, llm_endpoint, llm_model, *args):
    logging.info(f"LLM 类型: {llm_type}")
    logging.info(f"LLM 端点: {llm_endpoint}")
    logging.info(f"LLM 模型: {llm_model}")
    logging.info(f"=== 开始生成提示词（无优化） ===")

    max_attempts = config.MAX_ATTEMPTS
    for attempt in range(max_attempts):
        preprocess_prompt = f"""
        Analyse the following image description written in Chinese:
        "{description}"
        
        1. Translate the description into English.
        2. Determine if it is a complete scene description or just some scattered keywords: 
            - If it is a complete scene description, expand it with appropriate details.
            - Otherwise, generate a more detailed description based on the given keywords and create a complete scene description.
        
        3. Only generate final description of the scene as your output. The entire description should be quoted between '<<' and '>>'.
        """

        generated_description = call_llm(api_key, llm_type, llm_endpoint, llm_model, preprocess_prompt)
        logging.info(f"生成的描述： {generated_description}")

        processed_description = extract_content(generated_description, '<<', '>>')
        logging.info(f"清洗后的描述： {processed_description}")

        if processed_description:
            logging.info(f"成功生成提示词: {processed_description[:50]}...")
            return processed_description
        else:
            logging.warning(f"尝试 {attempt + 1}/{max_attempts}: 无法生成有效提示词，正在重新尝试。")

    logging.error("达到最大尝试次数，无法生成有效提示词，返回原始描述")
    return description

def generate_prompt_with_optimization(api_key, description, llm_type, llm_endpoint, llm_model, scene_type, *args):
    logging.info(f"LLM 类型: {llm_type}")
    logging.info(f"LLM 端点: {llm_endpoint}")
    logging.info(f"LLM 模型: {llm_model}")
    logging.info(f"场景类型: {scene_type}")
    logging.info(f"=== 开始生成提示词（使用优化） ===")

    max_attempts = config.MAX_ATTEMPTS
    for attempt in range(max_attempts):
        # 首先生成基础提示词
        base_prompt = generate_prompt_without_optimization(api_key, description, llm_type, llm_endpoint, llm_model, *args)
        
        if base_prompt == description:
            logging.warning(f"尝试 {attempt + 1}/{max_attempts}: 基础提示词生成失败，正在重新尝试。")
            continue

        # 使用模板增强生成的提示词
        enhanced_prompt = enhance_prompt_with_template(api_key, llm_type, llm_endpoint, llm_model, scene_type, description, base_prompt, base_prompt)

        if enhanced_prompt:
            logging.info(f"成功生成增强的提示词: {enhanced_prompt[:50]}...")
            return enhanced_prompt
        else:
            logging.warning(f"尝试 {attempt + 1}/{max_attempts}: 增强提示词失败，正在重新尝试。")

    logging.error("达到最大尝试次数，无法生成有效的增强提示词，返回原始述")
    return description

def generate_prompt(api_key, description, llm_type, llm_endpoint, llm_model, use_optimization, scene_type, *args):
    if use_optimization:
        prompt = generate_prompt_with_optimization(api_key, description, llm_type, llm_endpoint, llm_model, scene_type, *args)
    else:
        prompt = generate_prompt_without_optimization(api_key, description, llm_type, llm_endpoint, llm_model, *args)
    
    # 添加用户选择的参数和权重
    user_params = {}
    for i, options in enumerate(PARAMETER_TEMPLATES.values()):
        for j, option in enumerate(options):
            checkbox_value = args[i * 8 + j * 2]
            slider_value = args[i * 8 + j * 2 + 1]
            if checkbox_value and slider_value > 0:
                user_params[option] = int(slider_value)
    
    # 将参数附加到提示词后面
    for param, weight in user_params.items():
        param_formatted = param.lower().replace(" ", "_")
        prompt += f" {param_formatted}::{weight},"
    
    # 移除最后一个逗号（如果存在）
    prompt = prompt.rstrip(',')
    
    logging.info(f"最终生成的提示词（包含参数）: {prompt}")
    return prompt

# def extract_content(text, start_symbol, end_symbol):
#     try:
#         start_index = text.find(start_symbol)
#         if start_index != -1:
#             start_index += len(start_symbol)
#             end_index = text.rfind(end_symbol, start_index)
#             if end_index != -1:
#                 content = text[start_index:end_index]
#                 return content
#             else:
#                 raise ValueError(f"无法找到结束符号 {end_symbol}")
#         else:
#             raise ValueError(f"无法找到起始号 {start_symbol}")
#     except ValueError as e:
#         logging.error(f"提取内容失败: {e}")

def extract_content(text, start_symbol, end_symbol):
    pattern = re.escape(start_symbol) + '(.*?)' + re.escape(end_symbol)
    matches = re.findall(pattern, text)
    return matches[0] if matches else ""

def call_llm(api_key, llm_type, llm_endpoint, llm_model, prompt):
    if llm_type == 'openai':
        client = OpenAI(api_key=api_key, base_url=llm_endpoint)
        try:
            response = client.chat.completions.create(
                model=llm_model,
                messages=[
                    {"role": "system", "content": "You're a professional image prompt generation assistant, skilled in creating detailed and imaginative image descriptions."},
                    {"role": "user", "content": prompt}
                ]
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            logging.error(f"API 调用错误：{e}")
            return f"生成提示词时发生错误：{e}"
    
    elif llm_type == 'ollama':
        client = OpenAI(api_key='ollama', base_url=llm_endpoint)
        try:
            response = client.chat.completions.create(
                model=llm_model,
                messages=[
                    {"role": "system", "content": "You're a professional image prompt generation assistant, skilled in creating detailed and imaginative image descriptions."},
                    {"role": "user", "content": prompt}
                ]
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            logging.error(f"Ollama API 调用错误：{e}")
            return f"生成提示词时发生错误：{e}"
    
    else:
        return f"不支持的 LLM 类型：{llm_type}"

def parse_preprocessed_result(result):
    lines = result.split('\n')
    completeness = "Incomplete"
    processed_description = ""
    
    for line in lines:
        if line.startswith("Integraty"):
            completeness = line.split(":")[1].strip()
        elif line.startswith("处理后的描述:"):
            processed_description = line.split(":")[1].strip()
    
    return completeness, processed_description

def generate_image(prompt, workflow_name):
    max_retries = 3
    for attempt in range(max_retries):
        try:
            logging.info(f"尝试生成图像 (尝试 {attempt + 1}/{max_retries})")
            logging.info(f"使用提示词: {prompt[:100]}...")  # 只记录前100个字符
            logging.info(f"使用工作流: {workflow_name}")
            
            image_filename = comfyui_api.generate_image(prompt, workflow_name)
            if image_filename:
                image_path = os.path.join(config.OUTPUT_DIR, image_filename)
                if os.path.exists(image_path):
                    logging.info(f"成功生成图像: {image_path}")
                    return Image.open(image_path)
                else:
                    logging.error(f"生成的图像文件不存在: {image_path}")
            else:
                logging.error("ComfyUI API 未返回图像文件名")
            
            # 如果到达这里，说明当前尝试失败，等待后重试
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 5  # 递增等待时间
                logging.warning(f"图像生成失败，等待 {wait_time} 秒后重试")
                time.sleep(wait_time)
        except requests.exceptions.Timeout:
            logging.error("图像生成请求超时")
        except Exception as e:
            logging.error(f"图像生成错误: {str(e)}")
        
        if attempt == max_retries - 1:
            logging.error("达到最大重试次数，图像生成失败")
    
    return None

def process_excel(file_path, api_key, llm_type, llm_endpoint, llm_model, workflow_name, scene_type, use_optimization, progress=gr.Progress(), *args):
    try:
        logging.info(f"处理Excel文件: {file_path}")
        df = pd.read_excel(file_path)
        keyword_column = next((col for col in df.columns if col.lower() in ['关键词', 'keywords']), None)
        
        if keyword_column is None:
            return pd.DataFrame(), "Excel文件中未找到'关键词'或'Keywords'列。"
        
        excel_name = os.path.splitext(os.path.basename(file_path))[0]
        output_folder = os.path.join(config.OUTPUT_DIR, excel_name)
        os.makedirs(output_folder, exist_ok=True)
        
        total_rows = len(df[keyword_column])
        
        # 第一阶段：批量生成提示词
        progress(0, desc="开始生成提示词")
        prompts = []
        for index, keyword in enumerate(df[keyword_column]):
            prompt = generate_prompt(api_key, keyword, llm_type, llm_endpoint, llm_model, use_optimization, scene_type, *args)
            prompts.append(prompt)
            progress((index + 1) / total_rows / 2, desc=f"生成提示词进度: {index+1}/{total_rows}")
        
        # 第二阶段：批量生成图片
        progress(0.5, desc="开始生成图片")
        results = []
        for index, (keyword, prompt) in enumerate(zip(df[keyword_column], prompts)):
            logging.info(f"正在处理第 {index+1}/{total_rows} 个图像")
            
            # 尝试清理内存
            gc.collect()
            
            # 如果是 CUDA 环境，可以尝试清理 CUDA 缓存
            try:
                import torch
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                    logging.info("CUDA 缓存已清理")
            except ImportError:
                logging.info("torch 模块不可用，跳过 CUDA 缓存清理")
            
            # 暂停一小段时间，给系统一些时间来释放资源
            time.sleep(2)
            
            image = generate_image(prompt, workflow_name)
            if image:
                image_filename = f"{index+1:03d}_{keyword}.png"
                image_path = os.path.join(output_folder, image_filename)
                image.save(image_path)
                results.append((keyword, prompt, image_path))
                logging.info(f"成功生成并保存图像: {image_path}")
            else:
                results.append((keyword, prompt, "图像生成失败"))
                logging.error(f"无法为关键词 '{keyword}' 生成图像")
            
            progress(0.5 + (index + 1) / total_rows / 2, desc=f"生成图片进度: {index+1}/{total_rows}")
            
            # 每处理 5 张图片后，暂停一段时间
            if (index + 1) % 5 == 0:
                logging.info("暂停处理，等待资源释放...")
                time.sleep(5)  # 暂停 5 秒
        
        progress(1.0, desc="处理完成")
        return pd.DataFrame(results, columns=["关键词", "生成的提示词", "图片路径"]), "全部内容已生成"
    except Exception as e:
        logging.error(f"处理Excel文件时发生错误: {str(e)}")
        return pd.DataFrame(), f"处理Excel文件时发生错误: {str(e)}"

def export_results_to_excel(results, original_file_path, include_thumbnails):
    try:
        excel_name = os.path.splitext(os.path.basename(original_file_path))[0]
        output_file = os.path.join(config.OUTPUT_DIR, f"{excel_name}_results.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "处理结果"
        
        headers = ["关键词", "生成的提示词", "图片路径"]
        ws.append(headers)
        
        for _, row in results.iterrows():
            keyword, prompt, image_path = row
            ws.append([keyword, prompt, image_path])
            
            if include_thumbnails and os.path.exists(image_path):
                img = XLImage(image_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, ws.cell(row=ws.max_row, column=4).coordinate)
        
        wb.save(output_file)
        return output_file
    except Exception as e:
        logging.error(f"导出Excel结果时发生错误: {e}")
        return None

def create_interface():
    # 在函数开始处定义 custom_css
    custom_css = """
        .custom-textbox-output textarea {
            font-family: monospace;
            white-space: pre;
            overflow-x: auto;
            border: 1px solid #ddd;
            border-radius: 4px;
            padding: 10px;
            background-color: #f9f9f9;
        }
    """

    with gr.Blocks(css=custom_css) as interface:
        gr.Markdown("# 提示词生成器和图像生成器")
        
        with gr.Row():
            # 左侧栏
            with gr.Column(scale=1):
                api_key_input = gr.Textbox(label="API 密钥", type="password", value=config.OPENAI_API_KEY)
                save_api_key = gr.Checkbox(label="保存 API 密钥")
                
                llm_type = gr.Dropdown(choices=["openai", "ollama"], label="LLM 类型", value=config.LLM_TYPE)
                
                # OpenAI 设置
                with gr.Group(visible=llm_type.value == "openai") as openai_group:
                    openai_endpoint = gr.Textbox(
                        label="OpenAI Endpoint", 
                        value=config.OPENAI_ENDPOINT,
                        type="text",
                        placeholder="输入自定义endpoint或保留默认值",
                        interactive=True
                    )                   
                    openai_model = gr.Dropdown(
                        choices=["gpt-3.5-turbo", "gpt-4", "gpt-4o", "gpt-4o-mini", "gpt-4-turbo"],
                        label="OpenAI 模型",
                        value=config.OPENAI_MODEL,
                        interactive=True
                    )
                
                # Ollama 设置
                with gr.Group(visible=llm_type.value == "ollama") as ollama_group:
                    ollama_endpoint = gr.Textbox(label="Ollama Endpoint", value=config.OLLAMA_ENDPOINT, type="text", interactive=True)
                    ollama_model = gr.Dropdown(
                        choices=["llama3.1", "llama3"], 
                        label="Ollama 模型", 
                        value=config.OLLAMA_MODEL,
                        interactive=True
                    )

                description_input = gr.Textbox(label="请输入图像描述", lines=5)
                
                with gr.Row():
                    use_prompt_optimization = gr.Checkbox(label="使用提示词优化", value=True)
                    scene_type = gr.Dropdown(choices=SCENE_TYPES, label="场景类型", value=SCENE_TYPES[0])
                
                update_templates_button = gr.Button("更新提示词模板分析")
                
                generate_button = gr.Button("生成提示词", variant="primary")

                with gr.Row():
                    workflows = comfyui_api.list_workflows()
                    workflow_dropdown = gr.Dropdown(choices=workflows, label="选择工作流")
                    generate_image_button = gr.Button("生成图像")
                
                excel_file = gr.File(label="上传Excel文件(可选)")
                process_excel_button = gr.Button("批量处理Excel")
                
                excel_output = gr.Dataframe(
                    headers=["关键词", "生成的提示词", "图片路径"],
                    label="Excel处理结果"
                )
                include_thumbnails = gr.Checkbox(label="在Excel中包含缩略图", value=True)
                export_button = gr.Button("导出结果")
            
            # 右侧栏（用于显示生成的提示词和图像）
            with gr.Column(scale=1):
                with gr.Row():
                    output = gr.Textbox(label="生成的提示词", lines=10)

                with gr.Row():
                    output_image = gr.Image(label="生成的图像", type="pil")
                
                # 修改这里：添加样式和滚动条
                process_status = gr.Textbox(
                    label="当前处理状态",
                    lines=10,
                    max_lines=20,
                    elem_classes="custom-textbox-output"
                )

        # 下方参数选择区域（分为三列）
        with gr.Row():
            parameter_inputs = []
            for i, (category, options) in enumerate(PARAMETER_TEMPLATES.items()):
                with gr.Column(scale=1):
                    gr.Markdown(f"## {category}")
                    for option in options:
                        with gr.Row():
                            checkbox = gr.Checkbox(label=option)
                            slider = gr.Slider(minimum=0, maximum=10, step=1, value=0, label="权重")
                            parameter_inputs.extend([checkbox, slider])
                if (i + 1) % 3 == 0 and i < len(PARAMETER_TEMPLATES) - 1:
                    with gr.Row():
                        pass  # 创建新的行

        # 更新LLM设置
        def update_llm_settings(llm_type):
            is_openai = llm_type == "openai"
            is_ollama = llm_type == "ollama"
            return {
                openai_group: gr.update(visible=is_openai),
                ollama_group: gr.update(visible=is_ollama),
                openai_endpoint: gr.update(visible=is_openai),
                openai_model: gr.update(visible=is_openai),
                ollama_endpoint: gr.update(visible=is_ollama),
                ollama_model: gr.update(visible=is_ollama)
            }

        llm_type.change(
            update_llm_settings,
            inputs=[llm_type],
            outputs=[openai_group, ollama_group, openai_endpoint, openai_model, ollama_endpoint, ollama_model]
        )

        def get_current_llm_params(llm_type, openai_endpoint, openai_model, ollama_endpoint, ollama_model):
            if llm_type == "openai":
                return openai_endpoint, openai_model
            elif llm_type == "ollama":
                return ollama_endpoint, ollama_model
            else:
                raise ValueError(f"Unsupported LLM type: {llm_type}")

        def generate_prompt_wrapper(api_key, description, llm_type, openai_endpoint, openai_model, ollama_endpoint, ollama_model, use_optimization, scene_type, *args):
            endpoint, model = get_current_llm_params(llm_type, openai_endpoint, openai_model, ollama_endpoint, ollama_model)
            return generate_prompt(api_key, description, llm_type, endpoint, model, use_optimization, scene_type, *args)

        def process_excel_wrapper(file_path, api_key, llm_type, openai_endpoint, openai_model, ollama_endpoint, ollama_model, workflow_name, scene_type, use_optimization, progress=gr.Progress(), *args):
            endpoint, model = get_current_llm_params(llm_type, openai_endpoint, openai_model, ollama_endpoint, ollama_model)
            try:
                result = process_excel(file_path, api_key, llm_type, endpoint, model, workflow_name, scene_type, use_optimization, progress, *args)
                logging.info("process_excel completed successfully")
                return result
            except Exception as e:
                logging.error(f"Error in process_excel: {str(e)}")
                raise

        def update_templates_analysis(api_key, llm_type, openai_endpoint, openai_model, ollama_endpoint, ollama_model):
            endpoint, model = get_current_llm_params(llm_type, openai_endpoint, openai_model, ollama_endpoint, ollama_model)
            analysis_results = analyze_prompt_templates(api_key, llm_type, endpoint, model)
            # 这里可以选择将分析结果保存到文件或数据库中
            return "提示词模板分析已更新"

        generate_button.click(
            generate_prompt_wrapper,
            inputs=[
                api_key_input, 
                description_input, 
                llm_type,
                openai_endpoint,
                openai_model,
                ollama_endpoint,
                ollama_model,
                use_prompt_optimization,
                scene_type
            ] + parameter_inputs,
            outputs=output
        )

        generate_image_button.click(
            generate_image,
            inputs=[output, workflow_dropdown],
            outputs=output_image
        )

        process_excel_button.click(
            process_excel_wrapper,
            inputs=[
                excel_file, 
                api_key_input, 
                llm_type,
                openai_endpoint,
                openai_model,
                ollama_endpoint,
                ollama_model,
                workflow_dropdown,
                scene_type,
                use_prompt_optimization
            ] + parameter_inputs,
            outputs=[excel_output, process_status]
        )

        export_button.click(
            export_results_to_excel,
            inputs=[excel_output, excel_file, include_thumbnails],
            outputs=process_status
        )

        update_templates_button.click(
            update_templates_analysis,
            inputs=[
                api_key_input,
                llm_type,
                openai_endpoint,
                openai_model,
                ollama_endpoint,
                ollama_model
            ],
            outputs=process_status
        )

        # 添加场景类型选择的必填验证
        scene_type.change(lambda x: gr.update(variant="primary" if x else "secondary"), inputs=[scene_type], outputs=[generate_button])

        def update_status():
            status = ""
            while True:
                try:
                    log_entry = log_queue.get_nowait()
                    status += log_entry + "\n"
                except queue.Empty:
                    break
            return status

        interface.load(update_status, outputs=[process_status], every=1)

    return interface

def launch_prompt_generator():
    interface = create_interface()
    interface.queue()
    interface.launch(server_name="0.0.0.0", share=False)

if __name__ == "__main__":
    print("正在启动提示词生成器界面...")
    launch_prompt_generator()

